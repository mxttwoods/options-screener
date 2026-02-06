"""
Excel report generator.

Produces a multi-sheet .xlsx workbook styled for institutional review.
Deps: openpyxl (only external beyond pandas).
"""

from datetime import date
from typing import Optional

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F3864") if HAS_OPENPYXL else None
_HEADER_FONT = (
    Font(name="Calibri", bold=True, color="FFFFFF", size=10) if HAS_OPENPYXL else None
)
_BODY_FONT = Font(name="Calibri", size=10) if HAS_OPENPYXL else None
_THIN_BORDER = (
    Border(
        bottom=Side(style="thin", color="D9D9D9"),
    )
    if HAS_OPENPYXL
    else None
)
_CENTER = Alignment(horizontal="center", vertical="center") if HAS_OPENPYXL else None
_LEFT = Alignment(horizontal="left", vertical="center") if HAS_OPENPYXL else None

_GRADE_FILLS = (
    {
        "A+": "15803D",
        "A": "22C55E",
        "B": "EAB308",
        "C": "F97316",
        "D": "9CA3AF",
        "F": "DC2626",
    }
    if HAS_OPENPYXL
    else {}
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _auto_width(ws, min_width: int = 10, max_width: int = 30):
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = best


def _apply_body_style(ws, nrows: int, ncols: int):
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER


def _format_col(ws, col_idx: int, fmt: str, start_row: int = 2, end_row: int = 1000):
    """Apply a number format to a column."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            cell.number_format = fmt


def _write_sheet(wb, name: str, df: pd.DataFrame, col_formats: Optional[dict] = None):
    """Write a DataFrame as a styled sheet."""
    ws = wb.create_sheet(title=name)

    # Header row
    for c, header in enumerate(df.columns, 1):
        ws.cell(row=1, column=c, value=header)
    _style_header(ws, len(df.columns))

    # Data rows
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, float):
                cell.value = round(val, 6)
            else:
                cell.value = val

    nrows = len(df)
    ncols = len(df.columns)
    _apply_body_style(ws, nrows, ncols)

    # Number formats
    if col_formats:
        for col_name, fmt in col_formats.items():
            if col_name in df.columns:
                col_idx = list(df.columns).index(col_name) + 1
                _format_col(ws, col_idx, fmt, 2, nrows + 1)

    _auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


# ---------------------------------------------------------------------------
# Grade fill coloring
# ---------------------------------------------------------------------------


def _color_grades(ws, grade_col_idx: int, nrows: int):
    for row in range(2, nrows + 2):
        cell = ws.cell(row=row, column=grade_col_idx)
        grade = str(cell.value).strip()
        if grade in _GRADE_FILLS:
            cell.fill = PatternFill("solid", fgColor=_GRADE_FILLS[grade])
            if grade in ("A+", "A", "F"):
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            else:
                cell.font = Font(name="Calibri", bold=True, color="000000", size=10)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_excel(df: pd.DataFrame, output_path: Optional[str] = None) -> str:
    """
    Write a multi-sheet Excel workbook from the scan DataFrame.

    Returns the path of the written file.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is required for Excel export.  pip install openpyxl"
        )

    if df.empty:
        raise ValueError("No data to export.")

    if output_path is None:
        output_path = f"alpha_scan_{date.today().isoformat()}.xlsx"

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # -----------------------------------------------------------------------
    # Sheet 1: Rankings
    # -----------------------------------------------------------------------
    rank_df = df[
        [
            "ticker",
            "name",
            "grade",
            "score",
            "fund_score",
            "iv_score",
            "signal",
            "spot",
            "sector",
            "trend",
            "event_flag",
        ]
    ].copy()
    rank_df.columns = [
        "Ticker",
        "Name",
        "Grade",
        "Score",
        "Fund",
        "IV",
        "Signal",
        "Spot",
        "Sector",
        "Trend",
        "Event",
    ]
    ws = _write_sheet(
        wb,
        "Rankings",
        rank_df,
        {
            "Score": "0.0",
            "Fund": "0.0",
            "IV": "0.0",
            "Spot": "$#,##0.00",
        },
    )
    grade_idx = list(rank_df.columns).index("Grade") + 1
    _color_grades(ws, grade_idx, len(rank_df))

    # -----------------------------------------------------------------------
    # Sheet 2: Volatility
    # -----------------------------------------------------------------------
    vol_df = df[
        [
            "ticker",
            "spot",
            "dte",
            "expiration",
            "atm_iv",
            "hv_30",
            "iv_hv",
            "iv_pctl",
            "term_slope",
        ]
    ].copy()
    vol_df.columns = [
        "Ticker",
        "Spot",
        "DTE",
        "Expiration",
        "ATM IV",
        "HV 30d",
        "IV/HV",
        "IV Pctl",
        "Term Slope",
    ]
    _write_sheet(
        wb,
        "Volatility",
        vol_df,
        {
            "Spot": "$#,##0.00",
            "ATM IV": "0.0%",
            "HV 30d": "0.0%",
            "IV/HV": "0.00",
            "IV Pctl": "0",
            "Term Slope": "0.0000",
        },
    )

    # -----------------------------------------------------------------------
    # Sheet 3: Greeks
    # -----------------------------------------------------------------------
    greeks_df = df[
        [
            "ticker",
            "spot",
            "atm_strike",
            "dte",
            "call_delta",
            "put_delta",
            "gamma",
            "call_theta",
            "put_theta",
            "vega",
            "prob_itm_call",
        ]
    ].copy()
    greeks_df.columns = [
        "Ticker",
        "Spot",
        "ATM Strike",
        "DTE",
        "Call Delta",
        "Put Delta",
        "Gamma",
        "Call Theta",
        "Put Theta",
        "Vega",
        "P(ITM) Call",
    ]
    _write_sheet(
        wb,
        "Greeks",
        greeks_df,
        {
            "Spot": "$#,##0.00",
            "ATM Strike": "$#,##0.00",
            "Call Delta": "0.00",
            "Put Delta": "0.00",
            "Gamma": "0.000000",
            "Call Theta": "0.00",
            "Put Theta": "0.00",
            "Vega": "0.00",
            "P(ITM) Call": "0.0%",
        },
    )

    # -----------------------------------------------------------------------
    # Sheet 4: Fundamentals
    # -----------------------------------------------------------------------
    fund_df = df[
        [
            "ticker",
            "name",
            "sector",
            "market_cap",
            "pe",
            "roe",
            "rev_growth",
            "profit_margin",
            "op_margin",
            "debt_to_equity",
            "current_ratio",
            "beta",
        ]
    ].copy()
    fund_df.columns = [
        "Ticker",
        "Name",
        "Sector",
        "Mkt Cap",
        "P/E",
        "ROE",
        "Rev Growth",
        "Profit Mgn",
        "Op Mgn",
        "D/E",
        "Current Ratio",
        "Beta",
    ]
    _write_sheet(
        wb,
        "Fundamentals",
        fund_df,
        {
            "Mkt Cap": "#,##0",
            "P/E": "0.0",
            "ROE": "0.0%",
            "Rev Growth": "0.0%",
            "Profit Mgn": "0.0%",
            "Op Mgn": "0.0%",
            "D/E": "0.0",
            "Current Ratio": "0.00",
            "Beta": "0.00",
        },
    )

    # -----------------------------------------------------------------------
    # Sheet 5: Events
    # -----------------------------------------------------------------------
    events_df = df[
        [
            "ticker",
            "next_earnings",
            "earnings_dte",
            "ex_div_date",
            "event_flag",
        ]
    ].copy()
    events_df.columns = [
        "Ticker",
        "Next Earnings",
        "Earn DTE",
        "Ex-Div Date",
        "Flag",
    ]
    _write_sheet(wb, "Events", events_df)

    # -----------------------------------------------------------------------
    # Sheet 6: Methodology
    # -----------------------------------------------------------------------
    from . import config as cfg

    method_data = [
        ["Section", "Detail"],
        ["Run Date", date.today().isoformat()],
        ["Universe", f"yfinance equity screener, top {cfg.MAX_TICKERS} by EOD volume"],
        ["Fundamental Weight", f"{cfg.FUNDAMENTAL_WEIGHT:.0%}"],
        ["IV Weight", f"{cfg.IV_WEIGHT:.0%}"],
        ["Target DTE", str(cfg.TARGET_DTE)],
        ["HV Window", f"{cfg.HV_WINDOW} days"],
        ["Risk-Free Rate", f"{cfg.RISK_FREE_RATE:.2%}"],
        [
            "Trend Filter",
            f"{'On' if cfg.TREND_FILTER else 'Off'}, min score {cfg.MIN_TREND_SCORE}",
        ],
        ["Earnings Buffer", f"{cfg.EARNINGS_BUFFER_DAYS} days"],
        ["Greeks Model", "Black-Scholes (European, no dividends)"],
        ["IV Percentile", "Rank of current ATM IV vs local sqlite cache"],
        ["", ""],
        ["Fund Weights", ""],
        *[[f"  {k}", f"{v:.0%}"] for k, v in cfg.FUND_WEIGHTS.items()],
        ["", ""],
        ["IV Weights", ""],
        *[[f"  {k}", f"{v:.0%}"] for k, v in cfg.IV_WEIGHTS.items()],
        ["", ""],
        ["Signal Logic", ""],
        ["  IV/HV >= 1.50", "IV rich -- sell premium"],
        ["  IV/HV >= 1.30 + uptrend", "Sell CC / CSP"],
        ["  IV/HV <= 0.80 + uptrend", "IV cheap -- buy calls / LEAPS"],
        ["  IV/HV <= 0.80", "IV cheap -- debit spreads"],
        ["  Otherwise", "Neutral"],
        ["", ""],
        ["Grade Scale", "A+ (90+), A (80+), B (70+), C (60+), D (50+), F (<50)"],
    ]
    ws_m = wb.create_sheet(title="Methodology")
    for r, row in enumerate(method_data, 1):
        for c, val in enumerate(row, 1):
            cell = ws_m.cell(row=r, column=c, value=val)
            if r == 1:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = _CENTER
            else:
                cell.font = _BODY_FONT
                cell.alignment = _LEFT
    _auto_width(ws_m)

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    wb.save(output_path)
    return output_path
